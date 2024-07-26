import os
import boto3
import pandas as pd
from dotenv import load_dotenv
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

access_key = os.getenv('AWS_ACCESS_KEY_ID')
secret_key = os.getenv('AWS_SECRET_ACCESS_KEY')
region = os.getenv('REGION')
resource_arn = os.getenv('RESOURCE_ARN')
queue_arn = os.getenv('QUEUE_ARN')

session = boto3.Session(
    aws_access_key_id=access_key,
    aws_secret_access_key=secret_key
)

connect = boto3.client('connect', region_name=region)

params = {
    "ResourceArn": resource_arn,
    "StartTime": datetime(2024, 6, 1),
    "EndTime": datetime(2024, 6, 23),
    "Interval": {
        "TimeZone": "UTC",
        "IntervalPeriod": "WEEK"
    },
    "Filters": [
        {
            "FilterKey": "QUEUE",
            "FilterValues": [queue_arn]
        }
    ],
    "Groupings": ["QUEUE"],
    "Metrics": [
        {"Name": "ABANDONMENT_RATE"},
        {"Name": "AGENT_ANSWER_RATE"},
        {"Name": "CONTACTS_ABANDONED"},
        {
            "Name": "SERVICE_LEVEL",
            "Threshold": [{"Comparison": "LT", "ThresholdValue": 21}]
        },
        {
            "Name": "CONTACTS_HANDLED",
            "MetricFilters": [{"MetricFilterKey": "INITIATION_METHOD", "MetricFilterValues": ["INBOUND"]}]
        },
        {
            "Name": "CONTACTS_HANDLED",
            "MetricFilters": [{"MetricFilterKey": "INITIATION_METHOD", "MetricFilterValues": ["OUTBOUND"]}]
        },
        {"Name": "CONTACTS_QUEUED"}
    ]
}

response = connect.get_metric_data_v2(**params)

metric_results = response.get('MetricResults', [])

data = []

for result in metric_results:
    queue_id = result['Dimensions'].get('QUEUE')
    queue_arn = result['Dimensions'].get('QUEUE_ARN')
    start_time = result['MetricInterval']['StartTime'].replace(tzinfo=None)
    end_time = result['MetricInterval']['EndTime'].replace(tzinfo=None)
    interval_string = f"{start_time.date()} to {end_time.date()}"
    days_in_interval = (end_time - start_time).days + 1 

    for collection in result['Collections']:
        metric_name = collection['Metric']['Name']
        metric_value = collection['Value']

        metric_filters = collection['Metric'].get('MetricFilters', [])
        if metric_filters:
            for filter in metric_filters:
                if filter['MetricFilterKey'] == 'INITIATION_METHOD':
                    if 'INBOUND' in filter['MetricFilterValues']:
                        metric_name += ' INBOUND'
                    elif 'OUTBOUND' in filter['MetricFilterValues']:
                        metric_name += ' OUTBOUND'

        data.append({
            "Metric Name": metric_name,
            "Metric Value": metric_value,
            "Interval": interval_string
        })

        if metric_name == "CONTACTS_QUEUED":
            avg_calls_day = metric_value / days_in_interval
            data.append({
                "Metric Name": "AVG_CALLS_PER_DAY",
                "Metric Value": avg_calls_day,
                "Interval": interval_string
            })

df = pd.DataFrame(data)

def custom_aggregate(group):
    avg_metrics = ["ABANDONMENT_RATE", "AGENT_ANSWER_RATE", "AVG_CALLS_PER_DAY", "SERVICE_LEVEL"]
    if group.name in avg_metrics:
        return group.mean()
    else:
        return group.sum()

grouped_df = df.groupby(['Metric Name', 'Interval']).aggregate(custom_aggregate).reset_index()

pivot_df = grouped_df.pivot(index='Metric Name', columns='Interval', values='Metric Value')

# Calculate totals based on metric name
totals = {}
for metric in pivot_df.index:
    if metric in ["ABANDONMENT_RATE", "AGENT_ANSWER_RATE", "AVG_CALLS_PER_DAY", "SERVICE_LEVEL"]:
        totals[metric] = pivot_df.loc[metric].mean()
    else:
        totals[metric] = pivot_df.loc[metric].sum()

pivot_df['Total'] = pd.Series(totals)

# Create Excel file with formatted output
current_date = datetime.now().strftime("%Y-%m-%d")
excel_file = f"June_{current_date}.xlsx"
pivot_df.to_excel(excel_file, index=True)

# Load workbook to apply formatting
wb = load_workbook(excel_file)
ws = wb.active

# Set column widths and apply formatting
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Apply border to all cells
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
for row in ws.iter_rows():
    for cell in row:
        cell.border = thin_border

# Apply header styles
header_font = Font(bold=True)
for cell in ws[1]:
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Left-align the first column
for cell in ws['A']:
    cell.alignment = Alignment(horizontal="left")

# Define percentage metrics for formatting
percentage_metrics = ["ABANDONMENT_RATE", "AGENT_ANSWER_RATE", "SERVICE_LEVEL"]

# Apply percentage formatting
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    metric_name = row[0].value
    if metric_name in percentage_metrics:
        for cell in row[1:]:
            cell.value = cell.value / 100  # Scale the value to be between 0 and 1
            cell.number_format = '0.00%'

# Format AVG_CALLS_PER_DAY to 2 decimal places
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[0].value == "AVG_CALLS_PER_DAY":
        for cell in row[1:]:
            cell.number_format = '0.00'

# Save formatted Excel file
wb.save(excel_file)

print(f"Excel file '{excel_file}' has been created and formatted.")
